import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation

# --- Préférences utilisateur pour le formatage (en-FR) ---
CURRENCY_SYMBOL = "€"
DECIMAL_SEPARATOR = ","
GROUPING_SEPARATOR = " " # Espace insécable, souvent affiché comme un espace normal

# Styles Openpyxl
def apply_styles(sheet, dataframe):
    # Register styles if not already registered
    if 'header_style' not in sheet.workbook.named_styles:
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True)
        header_style.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet.workbook.add_named_style(header_style)

    if 'currency_style' not in sheet.workbook.named_styles:
        currency_style = NamedStyle(name="currency_style")
        currency_style.number_format = f'#,##0{DECIMAL_SEPARATOR}00 {CURRENCY_SYMBOL}'
        currency_style.alignment = Alignment(horizontal="right", vertical="center")
        sheet.workbook.add_named_style(currency_style)

    if 'currency_m_style' not in sheet.workbook.named_styles:
        currency_m_style = NamedStyle(name="currency_m_style")
        # Format pour millions - peut nécessiter ajustement selon préférence
        # '#,##0.0," M€"' ou '# ##0,0\ "M€"'
        currency_m_style.number_format = f'#,##0{DECIMAL_SEPARATOR}00{GROUPING_SEPARATOR}" {CURRENCY_SYMBOL}M"'
        currency_m_style.alignment = Alignment(horizontal="right", vertical="center")
        sheet.workbook.add_named_style(currency_m_style)

    if 'percent_style' not in sheet.workbook.named_styles:
        percent_style = NamedStyle(name="percent_style")
        percent_style.number_format = f'0{DECIMAL_SEPARATOR}0%' # Format pourcentages sans décimales
        percent_style.alignment = Alignment(horizontal="right", vertical="center")
        sheet.workbook.add_named_style(percent_style)

    if 'percent_1d_style' not in sheet.workbook.named_styles:
        percent_1d_style = NamedStyle(name="percent_1d_style")
        percent_1d_style.number_format = f'0{DECIMAL_SEPARATOR}0%' # Format pourcentages avec 1 décimale
        percent_1d_style.alignment = Alignment(horizontal="right", vertical="center")
        sheet.workbook.add_named_style(percent_1d_style)

    if 'integer_style' not in sheet.workbook.named_styles:
        integer_style = NamedStyle(name="integer_style")
        integer_style.number_format = f'#,##0'
        integer_style.alignment = Alignment(horizontal="right", vertical="center")
        sheet.workbook.add_named_style(integer_style)

    if 'decimal_1d_style' not in sheet.workbook.named_styles:
        decimal_1d_style = NamedStyle(name="decimal_1d_style")
        decimal_1d_style.number_format = f'#,##0{DECIMAL_SEPARATOR}0'
        decimal_1d_style.alignment = Alignment(horizontal="right", vertical="center")
        sheet.workbook.add_named_style(decimal_1d_style)


    # Apply header style
    for col in range(1, len(dataframe.columns) + 1):
         sheet[get_column_letter(col) + '1'].style = 'header_style'
    if dataframe.index.name is not None:
         sheet['A1'].style = 'header_style' # Style for the index header


    # Apply specific styles based on row/column content heuristic (can be improved)
    for r_idx, row in enumerate(dataframe.itertuples()):
        excel_row = r_idx + 2 # Adjust for 1-based index and header row
        for c_idx, col_name in enumerate(dataframe.columns):
            excel_col = c_idx + 1 # Adjust for 1-based index
            cell = sheet[get_column_letter(excel_col) + str(excel_row)]
            value = row[c_idx + 1] # Access value from named tuple

            # Simple heuristic for formatting
            if isinstance(value, (int, np.integer)) and not isinstance(value, bool):
                 cell.style = 'integer_style'
            elif isinstance(value, (float, np.floating)):
                 # Check for percentages or specific currencies
                 if any(p in str(col_name).lower() for p in ['rate', 'churn', 'expansion', '%', 'margin', 'retention']):
                      cell.style = 'percent_1d_style'
                 elif '€' in str(col_name) or 'revenue' in str(col_name).lower() or 'cost' in str(col_name).lower() or 'expense' in str(col_name).lower() or 'profit' in str(col_name).lower() or 'ebitda' in str(col_name).lower() or 'cac' in str(col_name).lower() or 'ltv' in str(col_name).lower() or 'arpu' in str(col_name).lower():
                      cell.style = 'currency_style'
                 else: # Default float formatting
                      cell.style = 'decimal_1d_style' # Or 'currency_style' if most floats are monetary

    # Auto-adjust column width (approximatif)
    for col in range(1, len(dataframe.columns) + 1):
         sheet.column_dimensions[get_column_letter(col)].width = 18 # Default width
         # Attempt to auto-adjust based on header length
         header_len = len(str(dataframe.columns[col-1]))
         sheet.column_dimensions[get_column_letter(col)].width = max(sheet.column_dimensions[get_column_letter(col)].width, header_len * 1.2)

    if dataframe.index.name is not None:
         sheet.column_dimensions['A'].width = max(sheet.column_dimensions['A'].width, len(str(dataframe.index.name)) * 1.2)


# --- Hypothèses du Modèle (ONGLET INPUTS) ---
# Les valeurs par défaut sont basées sur le document THE ARCHITECT ECOSYSTEM
# Années de projection
N_YEARS = 3
N_MONTHS = N_YEARS * 12
model_months = range(1, N_MONTHS + 1)
model_years = [2025 + (m - 1) // 12 for m in model_months] # Année fiscale basée sur le mois

# 1. Hypothèses d'Acquisition & Croissance
# Utilisateurs en début de Modèle (Month 1)
initial_free_users = 1000 # Hypothèse: Nombre initial d'utilisateurs Free
initial_pro_users = 50 # Hypothèse: Nombre initial d'utilisateurs Pro
initial_team_users = 5 # Hypothèse: Nombre initial d'utilisateurs Team
initial_team_accounts = 1 # Hypothèse: Nombre initial d'accounts Team
initial_enterprise_users = 1 # Hypothèse: Nombre initial d'utilisateurs Enterprise
initial_enterprise_accounts = 1 # Hypothèse: Nombre initial d'accounts Enterprise
initial_lifetime_users = 1 # Hypothèse: Nombre initial d'utilisateurs Lifetime

# Nombre de nouveaux utilisateurs FREE par mois (point de départ)
# Le document parle de croissance mensuelle des TAUX, pas du nombre. Interprétation:
# Le nombre de NOUVEAUX utilisateurs acquis chaque mois croit à un certain taux.
# Mettons un nombre initial et un taux de croissance mensuel des NOUVEAUX utilisateurs.
new_free_users_month_1 = 500 # Hypothèse: Nouveaux utilisateurs Free au premier mois du modèle
monthly_new_user_growth_rate = { # Taux de croissance MOIS sur MOIS des NOUVEAUX utilisateurs Free
    1: 0.15, # Growth rate for M1 -> M2 acquisitions
    2: 0.15,
    3: 0.15,
    4: 0.15,
    5: 0.15,
    6: 0.15,
    7: 0.22, # Start of Year 2 growth rate
    8: 0.22,
    9: 0.22,
    10: 0.22,
    11: 0.22,
    12: 0.22,
    13: 0.22,
    14: 0.22,
    15: 0.22,
    16: 0.22,
    17: 0.22,
    18: 0.22,
    19: 0.30, # Start of Year 3 growth rate
    20: 0.30,
    21: 0.30,
    22: 0.30,
    23: 0.30,
    24: 0.30,
    25: 0.28, # Document's paid rate seems lower in Y3? Let's average the organic/paid trends
    26: 0.28,
    27: 0.28,
    28: 0.28,
    29: 0.28,
    30: 0.28,
    31: 0.28,
    32: 0.28,
    33: 0.28,
    34: 0.28,
    35: 0.28,
    36: 0.28,
} # Simplification: on utilise un taux de croissance unique pour les nouveaux free

# Taux de Conversion FREE -> PAID (appliqué à la base Active Free chaque mois)
# Document: "Free-to-Paid Conversion Rate", "Lifetime License Uptake", "Team Conversion (% of paid users)"...
# L'architecture est Free -> Pro OR Lifetime OR Team. Puis Paid Individuel (Pro/LT) -> Team. Puis Team -> Enterprise.
free_to_pro_monthly_rate = {m: 0.035/12 for m in model_months} # 3.5% annuel / 12 mois en Y1
free_to_lifetime_monthly_rate = {m: 0.08/12 for m in model_months} # 8% annuel / 12 mois en Y1
free_to_team_monthly_rate_direct = {m: 0.0 for m in model_months} # Hypothèse: Pas de conversion directe Free -> Team majeure, passe par Paid Individual ou Paid User (cf. doc)

# Le document dit "Team Conversion (% of paid users)". Est-ce Paid INDIVIDUAL -> Team ? Ou Total Paid -> Team?
# Interprétons comme Paid INDIVIDUAL (Pro + Lifetime) -> Team ACCOUNT
paid_individual_to_team_account_monthly_rate = {
    m: (0.05/12) if m <=12 else ((0.15/12) if m <= 24 else (0.25/12)) for m in model_months
}

# Le document dit "Enterprise Conversion (% of team accounts)".
team_account_to_enterprise_account_monthly_rate = {
    m: (0.02/12) if m <=12 else ((0.08/12) if m <= 24 else (0.15/12)) for m in model_months
}

# Hypothèses sur les comptes multi-utilisateurs
avg_users_per_team_account = {m: 8 if m <= 12 else (10 if m <= 24 else 12) for m in model_months} # Moyenne évolue avec la maturité
avg_users_per_enterprise_account = {m: 50 if m <= 12 else (80 if m <= 24 else 100) for m in model_months} # Moyenne évolue

# Taux de Churn Mensuel (À TWEAKER)
monthly_churn_rate = {
    'Pro': {m: 0.055 if m <= 12 else (0.042 if m <= 24 else 0.035) for m in model_months},
    'Team': {m: 0.038 if m <= 12 else (0.030 if m <= 24 else 0.022) for m in model_months},
    'Enterprise': {m: 0.015 if m <= 12 else (0.012 if m <= 24 else 0.010) for m in model_months},
    'Lifetime': {m: 0.01 for m in model_months}, # Hypothèse: Churn très bas pour LT, non spécifié dans le doc
    'Free': {m: 0.20 for m in model_months}, # Hypothèse: Churn Free élevé, non spécifié dans le doc
}

# Taux d'Expansion Mensuel (affecte le MRR des clients qui NE churnent PAS) (À TWEAKER)
# L'expansion se traduit par plus d'utilisateurs dans les comptes Team/Enterprise, ou upsell potentiel sur Pro (moins fréquent)
# Simplification: On applique un taux d'augmentation du revenu MOIS sur MOIS pour les utilisateurs RETENUS dans chaque tier
monthly_expansion_rate = {
    'Pro': {m: 0.005 for m in model_months}, # Peu d'expansion sur Pro (uniquement via cross-sell futurs ?)
    'Team': {m: 0.01 if m <= 12 else (0.015 if m <= 24 else 0.02) for m in model_months}, # Augmentation utilisateurs par compte, cross-sell
    'Enterprise': {m: 0.015 if m <= 12 else (0.02 if m <= 24 else 0.025) for m in model_months}, # Augmentation utilisateurs par compte, cross-sell, modules
    'Lifetime': {m: 0.005 for m in model_months}, # Peu d'expansion sur LT
} # Attention: NRR = (1 - Churn Rate Mensuel + Expansion Rate Mensuel)^12 - approx.
# NRR Annuel ~ (1 - Churn Annuel + Expansion Annuel) - approx.


# 2. Hypothèses de Tarification & Revenus
# Prix Mensuels (pour Pro et Team) et Annuels (pour Enterprise)
# Lifetime est un prix Unique
monthly_prices = {
    'Pro': {m: 39 for m in model_months},
    'Team': {m: 29 if m <= 24 else 35 for m in model_months}, # Prix par utilisateur
}
annual_prices = {
    'Enterprise': {m: 500 if m <= 12 else (550 if m <= 24 else 600) for m in model_months}, # Prix par utilisateur, payé annuellement
}
one_time_license_price = {m: 50 for m in model_months} # Prix Licence unique

# Marketplace Revenue (Calculé Annuellement pour simplifier)
marketplace_commission_rate = {m: 0.0 for m in range(1, 13)} | {m: 0.15 for m in range(13, 25)} | {m: 0.18 for m in range(25, 37)} # En %
avg_marketplace_purchase_per_user_per_year = {m: 0 for m in range(1, 13)} | {m: 10 for m in range(13, 25)} | {m: 25 for m in range(25, 37)} # En €/utilisateur payant/an


# 3. Hypothèses de Coûts (en % du Revenu Total)
# Ces pourcentages sont appliqués au Revenu *Mensuel* Total
cost_structure_pct = {
    'Engineering': {m: 0.45 if m <= 12 else (0.38 if m <= 24 else 0.32) for m in model_months},
    'Product & Design': {m: 0.20 if m <= 12 else (0.18 if m <= 24 else 0.15) for m in model_months},
    'Marketing & Community': {m: 0.30 if m <= 12 else (0.25 if m <= 24 else 0.20) for m in model_months},
    'Customer Success': {m: 0.15 if m <= 12 else (0.12 if m <= 24 else 0.10) for m in model_months},
    'Infrastructure': {m: 0.15 if m <= 12 else (0.12 if m <= 24 else 0.08) for m in model_months},
    'G&A': {m: 0.15 if m <= 12 else (0.12 if m <= 24 else 0.10) for m in model_months},
}
# Note: Les pourcentages de coûts donnés dans le document (>100% en Y1)
# semblent être une *allocation* des dépenses par catégorie, pas des coûts variables marginaux.
# Ils semblent aussi basés sur le revenu annuel.
# Interprétation alternative: Coûts Fixes + Coûts Variables. Le modèle actuel utilise % du revenu = variable.
# Utilisons les % du document appliqués au revenu MENSUEL pour simuler la structure.
# total_cost_pct = {m: sum(cost_structure_pct[cat][m] for cat in cost_structure_pct) for m in model_months} # Vérification: >100% en Y1.
# Correction: Il faut probablement que ces pourcentages soient appliqués à une *marge brute* ou que le revenu soit plus haut initialement.
# Ou alors, il s'agit de coûts fixes/semi-variables qui sont *exprimés* en % du revenu dans le pitch *après* calcul.
# Utilisons les pourcentages comme donnés, appliqués au revenu total mensuel, ce qui implique des coûts > revenus initialement, ce qui est normal.
# MAIS, le document a aussi un coût "Infrastructure". Utiliser % du revenu pour Infrastructure est OK pour le hosting, mais pas pour l'équipe.
# Les autres catégories (Eng, Product, Mktg, CS, G&A) sont des coûts de personnel/fixes/variables qui ne DÉPENDENT pas directement du revenu.
# C'est une simplification du document. Pour un modèle plus précis, il faudrait input les coûts en dur ou par employé.
# Restons fidèles au document: coûts = % * Revenu Total Mensuel.

# Coût des biens vendus (COGS) - Séparé des OpEx
# COGS pourrait inclure des coûts de processing des licences uniques, ou fees marketplace
# Simplification: Mettons COGS à 0 ici, car non spécifié dans le document. L'infrastructure est dans OpEx.


# 4. Métriques de Performance
# Ces valeurs sont CALCULÉES par le modèle, mais le document donne des TARGETS. On peut les mettre ici comme référence.
target_metrics = {
    'End of Year 1': {'Total Free Users': 18500, 'Total Paid Users': 2300, 'Total Revenue (€)': 438640, 'Blended CAC (€)': 110, 'CAC Payback (months)': 7.5, 'LTV (€)': 415, 'LTV:CAC Ratio': 3.77, 'Gross Margin (%)': 0.75, 'Contribution Margin (%)': 0.45, 'Monthly Churn (%)': 0.055, 'Monthly Recurring Revenue (€)': 30470},
    'End of Year 2': {'Total Free Users': 62000, 'Total Paid Users': 13005, 'Total Revenue (€)': 2793904, 'Blended CAC (€)': 95, 'CAC Payback (months)': 5.2, 'LTV (€)': 615, 'LTV:CAC Ratio': 6.47, 'Gross Margin (%)': 0.80, 'Contribution Margin (%)': 0.55, 'Monthly Churn (%)': 0.042, 'Monthly Recurring Revenue (€)': 221050},
    'End of Year 3': {'Total Free Users': 185000, 'Total Paid Users': 57859, 'Total Revenue (€)': 17256613, 'Blended CAC (€)': 85, 'CAC Payback (months)': 3.8, 'LTV (€)': 1022, 'LTV:CAC Ratio': 12.02, 'Gross Margin (%)': 0.85, 'Contribution Margin (%)': 0.65, 'Monthly Churn (%)': 0.035, 'Monthly Recurring Revenue (€)': 1403401},
}
# Note: Gross Margin et Contribution Margin dans le document ne correspondent pas aux pourcentages de coûts donnés.
# GM = 1 - COGS%. CM = GM - Variable OpEx %. Ici, tous les OpEx sont des % de revenu.
# Nous allons CALCULER GM (basé sur COGS si non nul) et l'EBITDA Margin (basé sur Total OpEx).
# Les valeurs GM et CM du document seront des TARGETS.

# --- Modèle Mensuel ---

# Initialisation du DataFrame Mensuel
cols_monthly = [
    'Year', 'Month',
    'New Free Users', 'Active Free Users (Start)', 'Active Free Users (End)', 'Free Churn',
    'Converted to Pro', 'Converted to Lifetime', 'Converted to Team Accts (from Indiv Paid)', 'Converted to Enterprise Accts (from Team Accts)',
    'Active Pro Users (Start)', 'Active Pro Users (End)', 'Pro Churn', 'Pro Expansion Revenue',
    'Active Lifetime Users (Start)', 'Active Lifetime Users (End)', 'Lifetime Churn', 'Lifetime Expansion Revenue',
    'Active Team Users (Start)', 'Active Team Users (End)', 'Team Churn', 'Team Expansion Revenue',
    'Active Team Accounts (Start)', 'Active Team Accounts (End)', 'Team Acct Churn', 'Team Acct Expansion Revenue',
    'Active Enterprise Users (Start)', 'Active Enterprise Users (End)', 'Enterprise Churn', 'Enterprise Expansion Revenue',
    'Active Enterprise Accounts (Start)', 'Active Enterprise Accounts (End)', 'Enterprise Acct Churn', 'Enterprise Acct Expansion Revenue',
    'Total Paid Users (End)',
    'Revenue - Pro Subscriptions', 'Revenue - Team Subscriptions', 'Revenue - Enterprise Subscriptions',
    'Revenue - One-time Licenses', 'Revenue - Marketplace', 'Total Recurring Revenue (MRR)', 'Total Revenue (incl. One-time & MP)',
    'Cost - Engineering', 'Cost - Product & Design', 'Cost - Marketing & Community', 'Cost - Customer Success', 'Cost - Infrastructure', 'Cost - G&A',
    'Total OpEx',
    'Gross Profit', 'EBITDA'
]
df_monthly = pd.DataFrame(0.0, index=model_months, columns=cols_monthly)

# Peupler les valeurs initiales (Mois 1)
df_monthly.loc[1, 'Year'] = model_years[0]
df_monthly.loc[1, 'Month'] = 1
df_monthly.loc[1, 'Active Free Users (Start)'] = initial_free_users
df_monthly.loc[1, 'Active Pro Users (Start)'] = initial_pro_users
df_monthly.loc[1, 'Active Lifetime Users (Start)'] = initial_lifetime_users # Assuming this exists, if not, set initial_lifetime_users = 0
df_monthly.loc[1, 'Active Team Users (Start)'] = initial_team_users
df_monthly.loc[1, 'Active Enterprise Users (Start)'] = initial_enterprise_users
df_monthly.loc[1, 'Active Team Accounts (Start)'] = initial_team_users / avg_users_per_team_account[1] if avg_users_per_team_account[1] > 0 else 0 # Convert users to accounts
df_monthly.loc[1, 'Active Enterprise Accounts (Start)'] = initial_enterprise_users / avg_users_per_enterprise_account[1] if avg_users_per_enterprise_account[1] > 0 else 0 # Convert users to accounts

# S'assurer que les nombres d'accounts sont des entiers ou floats cohérents
df_monthly.loc[1, ['Active Team Accounts (Start)', 'Active Enterprise Accounts (Start)']] = df_monthly.loc[1, ['Active Team Accounts (Start)', 'Active Enterprise Accounts (Start)']].round(2)


# Boucle de simulation mensuelle
for m in model_months:
    year = model_years[m-1]
    df_monthly.loc[m, 'Year'] = year
    df_monthly.loc[m, 'Month'] = (m - 1) % 12 + 1

    # Définir les valeurs de départ (clients actifs en début de mois)
    if m > 1:
        df_monthly.loc[m, 'Active Free Users (Start)'] = df_monthly.loc[m-1, 'Active Free Users (End)']
        df_monthly.loc[m, 'Active Pro Users (Start)'] = df_monthly.loc[m-1, 'Active Pro Users (End)']
        df_monthly.loc[m, 'Active Lifetime Users (Start)'] = df_monthly.loc[m-1, 'Active Lifetime Users (End)']
        df_monthly.loc[m, 'Active Team Users (Start)'] = df_monthly.loc[m-1, 'Active Team Users (End)']
        df_monthly.loc[m, 'Active Enterprise Users (Start)'] = df_monthly.loc[m-1, 'Active Enterprise Users (End)']
        df_monthly.loc[m, 'Active Team Accounts (Start)'] = df_monthly.loc[m-1, 'Active Team Accounts (End)']
        df_monthly.loc[m, 'Active Enterprise Accounts (Start)'] = df_monthly.loc[m-1, 'Active Enterprise Accounts (End)']

    # 1. Acquisition de Nouveaux Utilisateurs Free
    if m == 1:
        new_free_users_this_month = new_free_users_month_1
    else:
        # Croissance basée sur le nombre de nouveaux utilisateurs du mois précédent
        new_free_users_this_month = df_monthly.loc[m-1, 'New Free Users'] * (1 + monthly_new_user_growth_rate.get(m-1, 0)) # Use m-1 rate for growth to month m

    df_monthly.loc[m, 'New Free Users'] = new_free_users_this_month.round() # Round new users to nearest integer

    # 2. Churn Utilisateurs Free (appliqué à la base de début de mois)
    active_free_start = df_monthly.loc[m, 'Active Free Users (Start)']
    free_churn_this_month = active_free_start * monthly_churn_rate['Free'][m]
    df_monthly.loc[m, 'Free Churn'] = free_churn_this_month.round()

    # 3. Conversions Free -> Paid (appliqué à la base de début de mois APRÈS Churn - ou pendant le mois?)
    # Simplification: Appliqué à la base de début de mois avant ajout des nouveaux et avant churn.
    # Alternative plus précise: Appliquer à la base Active Free moyenne du mois (Start + End) / 2
    # Gardons la simplification: Appliqué à la base Active Free (Start)
    converted_to_pro = active_free_start * free_to_pro_monthly_rate[m]
    converted_to_lifetime = active_free_start * free_to_lifetime_monthly_rate[m]
    # Direct Free -> Team est ignoré pour l'instant (si free_to_team_monthly_rate_direct est 0)

    df_monthly.loc[m, 'Converted to Pro'] = converted_to_pro.round()
    df_monthly.loc[m, 'Converted to Lifetime'] = converted_to_lifetime.round()
    # df_monthly.loc[m, 'Converted to Team Accts (from Free)'] = active_free_start * free_to_team_monthly_rate_direct[m] / avg_users_per_team_account[m] # If direct conversion is added


    # Base Free Active à la fin du mois = Start + New - Churn - Conversions
    df_monthly.loc[m, 'Active Free Users (End)'] = active_free_start + new_free_users_this_month - free_churn_this_month - converted_to_pro - converted_to_lifetime # - converted_to_team_direct

    # 4. Churn Utilisateurs Paid (appliqué à la base de début de mois)
    pro_churn = df_monthly.loc[m, 'Active Pro Users (Start)'] * monthly_churn_rate['Pro'][m]
    lifetime_churn = df_monthly.loc[m, 'Active Lifetime Users (Start)'] * monthly_churn_rate['Lifetime'][m]
    team_churn = df_monthly.loc[m, 'Active Team Users (Start)'] * monthly_churn_rate['Team'][m]
    enterprise_churn = df_monthly.loc[m, 'Active Enterprise Users (Start)'] * monthly_churn_rate['Enterprise'][m]

    # Churn des comptes Team et Enterprise (cohérent avec churn utilisateurs)
    team_acct_churn = df_monthly.loc[m, 'Active Team Accounts (Start)'] * monthly_churn_rate['Team'][m]
    enterprise_acct_churn = df_monthly.loc[m, 'Active Enterprise Accounts (Start)'] * monthly_churn_rate['Enterprise'][m]


    df_monthly.loc[m, 'Pro Churn'] = pro_churn.round()
    df_monthly.loc[m, 'Lifetime Churn'] = lifetime_churn.round()
    df_monthly.loc[m, 'Team Churn'] = team_churn.round()
    df_monthly.loc[m, 'Enterprise Churn'] = enterprise_churn.round()
    df_monthly.loc[m, 'Team Acct Churn'] = team_acct_churn.round(2)
    df_monthly.loc[m, 'Enterprise Acct Churn'] = enterprise_acct_churn.round(2)


    # 5. Conversions entre Tiers Paid (appliqué à la base de début de mois APRÈS Churn ou avant?)
    # Appliqué à la base de début de mois, avant churn
    # Conversion Paid Individual (Pro + Lifetime) -> Team Account
    active_paid_individual_start = df_monthly.loc[m, 'Active Pro Users (Start)'] + df_monthly.loc[m, 'Active Lifetime Users (Start)']
    converted_to_team_accts = active_paid_individual_start * paid_individual_to_team_account_monthly_rate[m]

    # Conversion Team Account -> Enterprise Account
    active_team_accounts_start = df_monthly.loc[m, 'Active Team Accounts (Start)']
    converted_to_enterprise_accts = active_team_accounts_start * team_account_to_enterprise_account_monthly_rate[m]

    df_monthly.loc[m, 'Converted to Team Accts (from Indiv Paid)'] = converted_to_team_accts.round(2)
    df_monthly.loc[m, 'Converted to Enterprise Accts (from Team Accts)'] = converted_to_enterprise_accts.round(2)

    # 6. Mise à jour des Bases Utilisateurs Actifs (End of Month)
    # La base End of Month est la base Start + Nouveaux de ce mois - Churn - Conversions SORTANTES + Conversions ENTRANTES
    # Pro: Start - Churn Pro + Conversions Free->Pro - Conversions Pro->Team (via Indiv Paid)
    df_monthly.loc[m, 'Active Pro Users (End)'] = df_monthly.loc[m, 'Active Pro Users (Start)'] - pro_churn + converted_to_pro - (converted_to_team_accts * (df_monthly.loc[m, 'Active Pro Users (Start)'] / active_paid_individual_start) if active_paid_individual_start > 0 else 0) # Approx: proportional split of indiv paid who convert

    # Lifetime: Start - Churn Lifetime + Conversions Free->Lifetime - Conversions LT->Team (via Indiv Paid)
    df_monthly.loc[m, 'Active Lifetime Users (End)'] = df_monthly.loc[m, 'Active Lifetime Users (Start)'] - lifetime_churn + converted_to_lifetime - (converted_to_team_accts * (df_monthly.loc[m, 'Active Lifetime Users (Start)'] / active_paid_individual_start) if active_paid_individual_start > 0 else 0) # Approx: proportional split

    # Team Accounts: Start - Churn Team Accts + Conversions Free->Team (Direct si activé) + Conversions Indiv Paid -> Team Accts - Conversions Team Accts -> Enterprise Accts
    df_monthly.loc[m, 'Active Team Accounts (End)'] = df_monthly.loc[m, 'Active Team Accounts (Start)'] - team_acct_churn + df_monthly.loc[m, 'Converted to Team Accts (from Indiv Paid)'] - converted_to_enterprise_accts
     # Ajouter la conversion Free -> Team directe si applicable
     # df_monthly.loc[m, 'Active Team Accounts (End)'] += df_monthly.loc[m, 'Converted to Team Accts (from Free)']

    # Enterprise Accounts: Start - Churn Enterprise Accts + Conversions Team Accts -> Enterprise Accts
    df_monthly.loc[m, 'Active Enterprise Accounts (End)'] = df_monthly.loc[m, 'Active Enterprise Accounts (Start)'] - enterprise_acct_churn + converted_to_enterprise_accts

    # Team Users: Active Accounts * Avg Users per Account (apply to End of Month Accounts)
    df_monthly.loc[m, 'Active Team Users (End)'] = df_monthly.loc[m, 'Active Team Accounts (End)'] * avg_users_per_team_account[m]

    # Enterprise Users: Active Accounts * Avg Users per Account (apply to End of Month Accounts)
    df_monthly.loc[m, 'Active Enterprise Users (End)'] = df_monthly.loc[m, 'Active Enterprise Accounts (End)'] * avg_users_per_enterprise_account[m]


    # Ensure user counts are integers at the end of the month calculation step (round only at the end of month)
    for col in ['Active Free Users', 'Active Pro Users', 'Active Lifetime Users', 'Active Team Users', 'Active Enterprise Users']:
         df_monthly.loc[m, col + ' (End)'] = df_monthly.loc[m, col + ' (End)'].round()

    # Ensure account counts are floats with limited decimals
    for col in ['Active Team Accounts', 'Active Enterprise Accounts']:
         df_monthly.loc[m, col + ' (End)'] = df_monthly.loc[m, col + ' (End)'].round(2)


    # Total Paid Users (End of Month)
    df_monthly.loc[m, 'Total Paid Users (End)'] = df_monthly.loc[m, 'Active Pro Users (End)'] + df_monthly.loc[m, 'Active Lifetime Users (End)'] + df_monthly.loc[m, 'Active Team Users (End)'] + df_monthly.loc[m, 'Active Enterprise Users (End)']


    # 7. Calcul des Revenus Mensuels Récurrents (MRR) par Tier
    # Revenu des utilisateurs retenus + Revenu des nouveaux clients ce mois dans ce tier
    # On applique le prix mensuel à la base d'utilisateurs *moyenne* du mois, ou à la base de *fin* de mois?
    # Standard SaaS: MRR = Nombre d'utilisateurs PAYANTS * ARPU mensuel
    # MRR par tier est calculé sur la base Active (End of Month)
    # L'expansion est gérée en augmentant l'ARPU des clients qui NE churnent PAS (complexe)
    # Simplification Expansion: On applique un taux d'expansion sur le REVENU généré par les clients RETENUS par rapport au mois précédent.
    # Plus simple: On calcule le MRR basé sur les utilisateurs fin de mois * leur ARPU mensuel actuel.
    # L'expansion se reflétera dans l'augmentation de l'ARPU au fil du temps dans les tiers Team/Enterprise (via avg_users_per_account)
    # L'ARPC est le prix annuel. ARPU mensuel = ARPC / 12.
    
    # Revenue - Pro Subscriptions
    # Nombre d'utilisateurs Pro * Prix Mensuel Pro
    # Utilise la moyenne des utilisateurs actifs sur le mois pour une meilleure précision ? (Start + End) / 2
    avg_pro_users_month = (df_monthly.loc[m, 'Active Pro Users (Start)'] + df_monthly.loc[m, 'Active Pro Users (End)']) / 2
    df_monthly.loc[m, 'Revenue - Pro Subscriptions'] = avg_pro_users_month * monthly_prices['Pro'][m] * (1 + monthly_expansion_rate['Pro'][m]) # Appliquer l'expansion

    # Revenue - Team Subscriptions (par utilisateur, payé mensuellement)
    avg_team_users_month = (df_monthly.loc[m, 'Active Team Users (Start)'] + df_monthly.loc[m, 'Active Team Users (End)']) / 2
    df_monthly.loc[m, 'Revenue - Team Subscriptions'] = avg_team_users_month * monthly_prices['Team'][m] * (1 + monthly_expansion_rate['Team'][m]) # Appliquer l'expansion

    # Revenue - Enterprise Subscriptions (par utilisateur, payé annuellement)
    # On doit convertir le prix annuel en revenu mensuel
    avg_enterprise_users_month = (df_monthly.loc[m, 'Active Enterprise Users (Start)'] + df_monthly.loc[m, 'Active Enterprise Users (End)']) / 2
    df_monthly.loc[m, 'Revenue - Enterprise Subscriptions'] = avg_enterprise_users_month * (annual_prices['Enterprise'][m] / 12) * (1 + monthly_expansion_rate['Enterprise'][m]) # Appliquer l'expansion

    # Revenue - One-time Licenses
    # Généré par les NOUVEAUX utilisateurs Free qui achètent la licence unique ce mois
    # Proportion de New Free Users qui prennent la Licence Unique = free_to_lifetime_monthly_rate * 12 (si on interprète le 8% comme annuel)
    # Let's assume the conversion to Lifetime user INCLUDES the purchase of the one-time license.
    # So, Revenue from One-time Licenses = Number of users converted to Lifetime * One-time price
    # This seems to contradict the doc which lists LT as 8% of FREE users AND one-time price.
    # Let's assume "One-time License Uptake" (8% of free users) is a conversion rate from Free users, separate from Pro/Team/Enterprise subscriptions, generating ONE-TIME revenue.
    # And "Lifetime License Uptake" in the User Growth Projection might mean users who convert to *any* paid tier and are expected to stay long term? Or a separate tier?
    # Given "Lifetime License Uptake: 8% of free users" AND "One-time License: €50", this seems like a SEPARATE flow.
    # Let's model: New Free Users -> X% buy One-time License (€50).
    # But the User Growth Projection lists "One-time License Users" and "Pro Subscription Users" separately. This means Lifetime License is a *tier*, not just a one-time purchase.
    # Let's reconcile:
    # Free -> Pro (Subscription)
    # Free -> Lifetime (Tier, pays One-Time Fee UPFRONT, then is tracked as a user with 0 MRR but tracked for churn/LTV)
    # Free -> Team (Subscription)
    # Paid Individual (Pro/LT) -> Team (Subscription)
    # Team -> Enterprise (Subscription)
    # This interpretation aligns the "One-time License Users" count with "Lifetime" users.
    # So, Revenue - One-time Licenses = (Number of users converted to Lifetime THIS MONTH) * Price.

    # Number of users converted to Lifetime = df_monthly.loc[m, 'Converted to Lifetime'] (this is users, not accounts)
    df_monthly.loc[m, 'Revenue - One-time Licenses'] = df_monthly.loc[m, 'Converted to Lifetime'] * one_time_license_price[m]

    # Revenue - Marketplace (Calculé Annuellement, on le mettra dans le résumé annuel)
    df_monthly.loc[m, 'Revenue - Marketplace'] = 0 # Calculé plus tard annuellement

    # Total Monthly Recurring Revenue (MRR)
    df_monthly.loc[m, 'Total Recurring Revenue (MRR)'] = df_monthly.loc[m, 'Revenue - Pro Subscriptions'] + df_monthly.loc[m, 'Revenue - Team Subscriptions'] + df_monthly.loc[m, 'Revenue - Enterprise Subscriptions'] # Exclut One-time and Marketplace

    # Total Revenue Mensuel (pour calcul des coûts)
    df_monthly.loc[m, 'Total Revenue (incl. One-time & MP)'] = df_monthly.loc[m, 'Total Recurring Revenue (MRR)'] + df_monthly.loc[m, 'Revenue - One-time Licenses'] + df_monthly.loc[m, 'Revenue - Marketplace']


    # 8. Calcul des Coûts Mensuels (basé sur % du Total Revenue Mensuel)
    total_monthly_revenue = df_monthly.loc[m, 'Total Revenue (incl. One-time & MP)']
    total_opex_this_month = 0
    for cost_cat in cost_structure_pct:
        cost_this_month = total_monthly_revenue * cost_structure_pct[cost_cat][m]
        df_monthly.loc[m, f'Cost - {cost_cat}'] = cost_this_month
        total_opex_this_month += cost_this_month

    df_monthly.loc[m, 'Total OpEx'] = total_opex_this_month

    # 9. Calcul du Profit
    # Gross Profit = Total Revenue - COGS. Assuming COGS is 0 as not specified.
    df_monthly.loc[m, 'Gross Profit'] = total_monthly_revenue # If COGS were included, subtract them here

    # EBITDA = Gross Profit - Total OpEx
    df_monthly.loc[m, 'EBITDA'] = df_monthly.loc[m, 'Gross Profit'] - total_opex_this_month


# --- Agrégation Annuelle et Métriques ---

# Initialisation du DataFrame Annuel
cols_annual = [f'FY{y}' for y in sorted(list(set(model_years)))]
rows_annual = [
    'Total Free Users (End Year)',
    'Total Paid Users (End Year)',
    'Paid User Mix: Pro (%)', 'Paid User Mix: Lifetime (%)', 'Paid User Mix: Team (%)', 'Paid User Mix: Enterprise (%)',
    'Total Revenue (€)',
    'YoY Revenue Growth (%)',
    'Revenue - Pro Subscriptions (€)', 'Revenue - Team Subscriptions (€)', 'Revenue - Enterprise Subscriptions (€)',
    'Revenue - One-time Licenses (€)', 'Revenue - Marketplace (€)',
    'Total Recurring Revenue (MRR End Year, €)', 'Total Recurring Revenue (MRR Avg Year, €)',
    'Total OpEx (€)',
    'EBITDA (€)',
    'Gross Margin (%)', 'EBITDA Margin (%)',
    'Blended CAC (€)',
    'CAC Payback (Months)',
    'Average Revenue Per Paid User (ARPU, €/year)',
    'Lifetime Value (LTV, €)',
    'LTV:CAC Ratio',
    'Monthly Churn Rate (Avg Weighted, %)',
    'Net Revenue Retention (NRR, %)',
]
df_annual = pd.DataFrame(0.0, index=rows_annual, columns=cols_annual)

for year in sorted(list(set(model_years))):
    # Filtrer les données pour l'année en cours
    df_year_months = df_monthly[df_monthly['Year'] == year]
    year_col = f'FY{year}'

    # Statistiques Utilisateurs
    df_annual.loc['Total Free Users (End Year)', year_col] = df_year_months['Active Free Users (End)'].iloc[-1]
    df_annual.loc['Total Paid Users (End Year)', year_col] = df_year_months['Total Paid Users (End)'].iloc[-1]

    # Mix Paid Users (basé sur fin d'année)
    total_paid_end_year = df_annual.loc['Total Paid Users (End Year)', year_col]
    if total_paid_end_year > 0:
        df_annual.loc['Paid User Mix: Pro (%)', year_col] = df_year_months['Active Pro Users (End)'].iloc[-1] / total_paid_end_year
        df_annual.loc['Paid User Mix: Lifetime (%)', year_col] = df_year_months['Active Lifetime Users (End)'].iloc[-1] / total_paid_end_year
        df_annual.loc['Paid User Mix: Team (%)', year_col] = df_year_months['Active Team Users (End)'].iloc[-1] / total_paid_end_year
        df_annual.loc['Paid User Mix: Enterprise (%)', year_col] = df_year_months['Active Enterprise Users (End)'].iloc[-1] / total_paid_end_year
    else:
         df_annual.loc[['Paid User Mix: Pro (%)', 'Paid User Mix: Lifetime (%)', 'Paid User Mix: Team (%)', 'Paid User Mix: Enterprise (%)'], year_col] = 0

    # Revenus Annuels
    df_annual.loc['Revenue - Pro Subscriptions (€)', year_col] = df_year_months['Revenue - Pro Subscriptions'].sum()
    df_annual.loc['Revenue - Team Subscriptions (€)', year_col] = df_year_months['Revenue - Team Subscriptions'].sum()
    df_annual.loc['Revenue - Enterprise Subscriptions (€)', year_col] = df_year_months['Revenue - Enterprise Subscriptions'].sum()
    df_annual.loc['Revenue - One-time Licenses (€)', year_col] = df_year_months['Revenue - One-time Licenses'].sum()

    # Revenu Marketplace (basé sur moyenne des utilisateurs payants actifs pendant l'année)
    # Avg Paid Users during the year (simplification: average of start and end of month users for all months in the year)
    avg_paid_users_year = df_year_months[['Active Pro Users (End)', 'Active Lifetime Users (End)', 'Active Team Users (End)', 'Active Enterprise Users (End)']].sum(axis=1).mean()
    # Marketplace revenue = Avg Paid Users * Avg Purchase per User per Year * Commission Rate
    # Commission rate and purchase per user per year might vary by month, use the average for the year or value at end of year?
    # Let's use the value at the end of the year for simplicity, multiplied by the average user base.
    last_month_of_year = df_year_months.index.max()
    mp_commission_rate_year = marketplace_commission_rate.get(last_month_of_year, 0)
    avg_mp_purchase_year = avg_marketplace_purchase_per_user_per_year.get(last_month_of_year, 0)
    df_annual.loc['Revenue - Marketplace (€)', year_col] = avg_paid_users_year * avg_mp_purchase_year * mp_commission_rate_year

    # Total Revenue Annuel
    df_annual.loc['Total Revenue (€)', year_col] = df_annual.loc['Revenue - Pro Subscriptions (€)', year_col] + df_annual.loc['Revenue - Team Subscriptions (€)', year_col] + df_annual.loc['Revenue - Enterprise Subscriptions (€)', year_col] + df_annual.loc['Revenue - One-time Licenses (€)', year_col] + df_annual.loc['Revenue - Marketplace (€)', year_col]

    # YoY Revenue Growth
    if year > model_years[0]:
        prev_year_col = f'FY{year-1}'
        prev_year_revenue = df_annual.loc['Total Revenue (€)', prev_year_col]
        if prev_year_revenue > 0:
            df_annual.loc['YoY Revenue Growth (%)', year_col] = (df_annual.loc['Total Revenue (€)', year_col] - prev_year_revenue) / prev_year_revenue
        else:
            df_annual.loc['YoY Revenue Growth (%)', year_col] = np.nan # Or a large number if growth from 0

    # MRR End Year
    df_annual.loc['Total Recurring Revenue (MRR End Year, €)', year_col] = df_year_months['Total Recurring Revenue (MRR)'].iloc[-1]
    # MRR Avg Year (average of monthly MRR values during the year)
    df_annual.loc['Total Recurring Revenue (MRR Avg Year, €)', year_col] = df_year_months['Total Recurring Revenue (MRR)'].mean()


    # Coûts Annuels
    df_annual.loc['Total OpEx (€)', year_col] = df_year_months['Total OpEx'].sum()

    # EBITDA Annuel
    df_annual.loc['EBITDA (€)', year_col] = df_year_months['EBITDA'].sum()

    # Marges
    df_annual.loc['Gross Margin (%)', year_col] = df_year_months['Gross Profit'].sum() / df_annual.loc['Total Revenue (€)', year_col] if df_annual.loc['Total Revenue (€)', year_col] > 0 else 0
    df_annual.loc['EBITDA Margin (%)', year_col] = df_annual.loc['EBITDA (€)', year_col] / df_annual.loc['Total Revenue (€)', year_col] if df_annual.loc['Total Revenue (€)', year_col] > 0 else 0


    # Métriques SaaS
    # Blended CAC (€) = Total S&M annuel / Nouveaux clients payants ANNUELS
    # Nouveaux clients payants ANNUELS = Nouveaux utilisateurs Pro + Nouveaux utilisateurs Lifetime + Nouveaux utilisateurs Team + Nouveaux utilisateurs Enterprise
    # Attention: Conversions entre tiers (Indiv->Team, Team->Enterprise) NE SONT PAS de nouveaux clients acquis, mais des upgrades.
    # Nouveaux clients payants sont ceux qui viennent directement de la base FREE ce mois-ci et qui CONVERTISSENT pour la PREMIERE fois.
    # Total Nouveaux clients payants cette année = Somme(Converted to Pro + Converted to Lifetime + Converted to Team Direct) sur les mois de l'année.
    # (On ignore ici les free->team direct si rate est 0)
    new_paying_customers_this_year = df_year_months['Converted to Pro'].sum() + df_year_months['Converted to Lifetime'].sum() + df_year_months['Converted to Team Accts (from Free)'].sum() * avg_users_per_team_account.get(last_month_of_year, 1) # Assuming direct conversion to Team accounts represents new paying users

    total_sm_opex_year = df_year_months['Cost - Marketing & Community'].sum() # Assuming M&C is primarily S&M
    df_annual.loc['Blended CAC (€)', year_col] = total_sm_opex_year / new_paying_customers_this_year if new_paying_customers_this_year > 0 else np.nan # Use np.nan for division by zero

    # CAC Payback (Months) = Blended CAC / (ARPU Mensuel Moyen * Marge Brute %)
    # ARPU Mensuel Moyen de l'année = Total Recurring Revenue (Avg Year) / Avg Paid Users during the year (approx)
    avg_recurring_revenue_month_year = df_annual.loc['Total Recurring Revenue (MRR Avg Year, €)', year_col] / 12 if df_annual.loc['Total Recurring Revenue (MRR Avg Year, €)', year_col] > 0 else 0
    gross_margin_pct_year = df_annual.loc['Gross Margin (%)', year_col]

    if avg_paid_users_year > 0:
        arpu_monthly_avg_year = avg_recurring_revenue_month_year / avg_paid_users_year
    else:
        arpu_monthly_avg_year = 0

    monthly_gross_profit_per_paid_user = arpu_monthly_avg_year * gross_margin_pct_year

    df_annual.loc['CAC Payback (Months)', year_col] = df_annual.loc['Blended CAC (€)', year_col] / monthly_gross_profit_per_paid_user if monthly_gross_profit_per_paid_user > 0 else np.nan


    # Average Revenue Per Paid User (ARPU, €/year)
    # Basé sur le revenu récurrent total de l'année / le nombre moyen d'utilisateurs payants actifs pendant l'année
    # Ou plus simple: Total Revenue / Paid Users End of Year (moins précis)
    # Utilisons MRR fin d'année * 12 / Paid Users fin d'année (approx ARR / Paid Users)
    total_paid_end_year = df_annual.loc['Total Paid Users (End Year)', year_col]
    arr_end_year = df_annual.loc['Total Recurring Revenue (MRR End Year, €)', year_col] * 12
    df_annual.loc['Average Revenue Per Paid User (ARPU, €/year)', year_col] = arr_end_year / total_paid_end_year if total_paid_end_year > 0 else np.nan


    # Lifetime Value (LTV, €) - Simplifié
    # LTV = ARPU Annuel * Marge Brute % / Taux de Churn Annuel
    # Taux de Churn Annuel (Moyen Pondéré) - Approx. 1 - (1 - Churn Mensuel Moyen Pondéré)^12
    # Calcul du Churn Mensuel Moyen Pondéré de l'année: Somme(Churn Mensuel Tier * Avg Active Users Tier) / Somme(Avg Active Users Tier) pour chaque mois de l'année, puis moyenne annuelle.
    # Simplification: Moyenne des taux de churn mensuels pondérée par la base d'utilisateurs de début de mois pour CHAQUE mois, puis moyenne annuelle.
    weighted_avg_monthly_churn_list = []
    weighted_avg_monthly_expansion_list = []
    weighted_avg_monthly_nrr_list = []

    for month_m in df_year_months.index:
         active_users_start_m = {
              'Pro': df_monthly.loc[month_m, 'Active Pro Users (Start)'],
              'Team': df_monthly.loc[month_m, 'Active Team Users (Start)'],
              'Enterprise': df_monthly.loc[month_m, 'Active Enterprise Users (Start)'],
              'Lifetime': df_monthly.loc[month_m, 'Active Lifetime Users (Start)'],
         }
         total_active_paid_start_m = sum(active_users_start_m.values())

         if total_active_paid_start_m > 0:
              weighted_churn_m = sum(active_users_start_m[tier] * monthly_churn_rate[tier][month_m] for tier in ['Pro', 'Team', 'Enterprise', 'Lifetime']) / total_active_paid_start_m
              weighted_expansion_m = sum(active_users_start_m[tier] * monthly_expansion_rate[tier][month_m] for tier in ['Pro', 'Team', 'Enterprise', 'Lifetime']) / total_active_paid_start_m
              # Approximation NRR mensuel pondéré: 1 - Churn + Expansion
              weighted_nrr_m = 1 - weighted_churn_m + weighted_expansion_m

         else:
              weighted_churn_m = np.nan # No paid users
              weighted_expansion_m = np.nan
              weighted_nrr_m = np.nan

         weighted_avg_monthly_churn_list.append(weighted_churn_m)
         weighted_avg_monthly_expansion_list.append(weighted_expansion_m)
         weighted_avg_monthly_nrr_list.append(weighted_nrr_m)

    # Moyenne annuelle des taux mensuels pondérés
    avg_weighted_monthly_churn_year = np.nanmean(weighted_avg_monthly_churn_list) # Use nanmean to ignore NaNs
    avg_weighted_monthly_expansion_year = np.nanmean(weighted_avg_monthly_expansion_list)
    avg_weighted_monthly_nrr_year = np.nanmean(weighted_avg_monthly_nrr_list)

    df_annual.loc['Monthly Churn Rate (Avg Weighted, %)', year_col] = avg_weighted_monthly_churn_year

    # LTV = ARPU Annuel * Marge Brute % / Taux de Churn Annuel (Approx)
    # Taux de Churn Annuel (Approx) = 1 - (1 - Churn Mensuel Moyen Pondéré Annuel)^12
    annual_churn_rate_approx = 1 - (1 - avg_weighted_monthly_churn_year)**12 if not np.isnan(avg_weighted_monthly_churn_year) else np.nan

    # Utilise l'ARPU annuel (basé sur revenu récurrent) pour le calcul LTV
    arpu_annual_year = df_annual.loc['Average Revenue Per Paid User (ARPU, €/year)', year_col]

    df_annual.loc['Lifetime Value (LTV, €)', year_col] = (arpu_annual_year * gross_margin_pct_year) / annual_churn_rate_approx if annual_churn_rate_approx > 0 else np.nan

    # LTV:CAC Ratio
    df_annual.loc['LTV:CAC Ratio', year_col] = df_annual.loc['Lifetime Value (LTV, €)', year_col] / df_annual.loc['Blended CAC (€)', year_col] if not np.isnan(df_annual.loc['Blended CAC (€)', year_col]) else np.nan

    # Net Revenue Retention (NRR, %) - Calcul Annuel
    # NRR = (Revenu Recurring de l'année N des clients qui existaient au début de l'année N) / (Revenu Recurring de l'année N-1 des clients qui existaient au début de l'année N)
    # This requires tracking revenue cohort by cohort annually, which is complex in this monthly model.
    # Simplification: Use the average weighted monthly NRR compounded annually
    # NRR Annuel = (Moyenne NRR Mensuel Annuel)^12 - This is mathematically incorrect.
    # Correct Approx NRR Annuel = 1 - Churn Annuel + Expansion Annuel (calculated from monthly rates)
    # Let's use the average weighted monthly NRR as a proxy for monthly NRR.
    # NRR Annuel (approx) = (1 - Avg Weighted Monthly Churn + Avg Weighted Monthly Expansion)^12
    # No, Expansion is applied to revenue. NRR reflects revenue change from retained customers.
    # NRR Annuel = Revenue from Retained Cohorts in Year N / Revenue from Same Cohorts in Year N-1
    # Let's calculate this accurately: Sum revenue from all cohorts older than Year N over months in Year N, divided by sum revenue from same cohorts over months in Year N-1.
    if year > model_years[0]:
         prev_year = year - 1
         revenue_from_old_cohorts_this_year = df_monthly[df_monthly['Year'] == year].loc[:, [col for col in df_monthly.columns if 'Revenue -' in col and 'Total' not in col and 'One-time' not in col and 'Marketplace' not in col]].sum().sum() # Sum of monthly recurring revenue from all cohorts THIS YEAR

         # Calculate revenue from cohorts *existing at start of Year N* during Year N-1
         revenue_from_start_of_year_cohort_prev_year = df_monthly[df_monthly['Year'] == prev_year].loc[:, [col for col in df_monthly.columns if 'Revenue -' in col and 'Total' not in col and 'One-time' not in col and 'Marketplace' not in col]].sum().sum() # Total recurring revenue in Year N-1

         # This is still not quite right. Need to isolate revenue *from the specific set of customers* that were active at the START of Year N.
         # This requires cohort tracking at a monthly revenue level, summing up.
         # Let's simplify for now: NRR = (1 - Avg Weighted Monthly Churn + Avg Weighted Monthly Expansion)
         # Use the average weighted monthly NRR as a proxy.
         df_annual.loc['Net Revenue Retention (NRR, %)', year_col] = avg_weighted_monthly_nrr_year # Using the average weighted monthly NRR directly as the input states %

    else:
         df_annual.loc['Net Revenue Retention (NRR, %)', year_col] = np.nan # NRR not applicable in Year 1


# --- Création du Fichier Excel ---
output_filename = 'architect_financial_model.xlsx'

with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    # --- Onglet Inputs ---
    df_inputs_data = {
        'Parameter': [], 'Value': [], 'Notes': []
    }

    def add_input(param, value, notes=''):
        df_inputs_data['Parameter'].append(param)
        df_inputs_data['Value'].append(value)
        df_inputs_data['Notes'].append(notes)

    add_input('Number of Projection Years', N_YEARS, 'Nombre d\'années simulées')
    add_input('--- Initial Users (Start of Month 1) ---', '', '')
    add_input('Initial Free Users', initial_free_users, 'Nombre d\'utilisateurs Free au début du modèle')
    add_input('Initial Pro Users', initial_pro_users, 'Nombre d\'utilisateurs Pro au début du modèle')
    add_input('Initial Lifetime Users', initial_lifetime_users, 'Nombre d\'utilisateurs Lifetime au début du modèle')
    add_input('Initial Team Users', initial_team_users, 'Nombre d\'utilisateurs Team au début du modèle')
    add_input('Initial Enterprise Users', initial_enterprise_users, 'Nombre d\'utilisateurs Enterprise au début du modèle')
    add_input('--- Acquisition & Growth ---', '', '')
    add_input('New Free Users Month 1', new_free_users_month_1, 'Nombre de nouveaux utilisateurs Free acquis le premier mois')
    add_input('Monthly New User Growth Rate - Y1 (%)', monthly_new_user_growth_rate[1]*12, 'Taux de croissance ANNUEL des NOUVEAUX utilisateurs Free en Y1 (converti du mensuel)') # Show annual rate
    add_input('Monthly New User Growth Rate - Y2 (%)', monthly_new_user_growth_rate[13]*12, 'Taux de croissance ANNUEL des NOUVEAUX utilisateurs Free en Y2') # Show annual rate
    add_input('Monthly New User Growth Rate - Y3 (%)', monthly_new_user_growth_rate[25]*12, 'Taux de croissance ANNUEL des NOUVEAUX utilisateurs Free en Y3') # Show annual rate
    # Note: Monthly growth rate is applied month over month compounding. The input here is the monthly rate, but displaying annual in Excel is clearer.

    add_input('--- Conversion Rates (Monthly %) ---', '', '')
    # Display annual rates in Inputs sheet for clarity, but use monthly in model
    add_input('Free to Pro (Annual %) - Y1', free_to_pro_monthly_rate[1]*12, 'Conversion Free vers Pro (sur base Free Active)')
    add_input('Free to Pro (Annual %) - Y2', free_to_pro_monthly_rate[13]*12, '')
    add_input('Free to Pro (Annual %) - Y3', free_to_pro_monthly_rate[25]*12, '')
    add_input('Free to Lifetime (Annual %) - Y1', free_to_lifetime_monthly_rate[1]*12, 'Conversion Free vers Lifetime (sur base Free Active)')
    add_input('Free to Lifetime (Annual %) - Y2', free_to_lifetime_monthly_rate[13]*12, '')
    add_input('Free to Lifetime (Annual %) - Y3', free_to_lifetime_monthly_rate[25]*12, '')
    add_input('Paid Individual to Team Acct (Annual %) - Y1', paid_individual_to_team_account_monthly_rate[1]*12, 'Conversion Pro/Lifetime vers Team Accounts (sur base Indiv Paid Active)')
    add_input('Paid Individual to Team Acct (Annual %) - Y2', paid_individual_to_team_account_monthly_rate[13]*12, '')
    add_input('Paid Individual to Team Acct (Annual %) - Y3', paid_individual_to_team_account_monthly_rate[25]*12, '')
    add_input('Team Acct to Enterprise Acct (Annual %) - Y1', team_account_to_enterprise_account_monthly_rate[1]*12, 'Conversion Team Accounts vers Enterprise Accounts (sur base Team Accounts Active)')
    add_input('Team Acct to Enterprise Acct (Annual %) - Y2', team_account_to_enterprise_account_monthly_rate[13]*12, '')
    add_input('Team Acct to Enterprise Acct (Annual %) - Y3', team_account_to_enterprise_account_monthly_rate[25]*12, '')

    add_input('--- Churn Rates (Monthly %) ---', '', '')
    add_input('Churn Rate Free (Monthly %)', monthly_churn_rate['Free'][1], '')
    add_input('Churn Rate Pro (Monthly %) - Y1', monthly_churn_rate['Pro'][1], '')
    add_input('Churn Rate Pro (Monthly %) - Y2', monthly_churn_rate['Pro'][13], '')
    add_input('Churn Rate Pro (Monthly %) - Y3', monthly_churn_rate['Pro'][25], '')
    add_input('Churn Rate Lifetime (Monthly %)', monthly_churn_rate['Lifetime'][1], 'Assumed low churn for Lifetime')
    add_input('Churn Rate Team (Monthly %) - Y1', monthly_churn_rate['Team'][1], '')
    add_input('Churn Rate Team (Monthly %) - Y2', monthly_churn_rate['Team'][13], '')
    add_input('Churn Rate Team (Monthly %) - Y3', monthly_churn_rate['Team'][25], '')
    add_input('Churn Rate Enterprise (Monthly %) - Y1', monthly_churn_rate['Enterprise'][1], '')
    add_input('Churn Rate Enterprise (Monthly %) - Y2', monthly_churn_rate['Enterprise'][13], '')
    add_input('Churn Rate Enterprise (Monthly %) - Y3', monthly_churn_rate['Enterprise'][25], '')

    add_input('--- Expansion Rates (Monthly % of Revenue) ---', '', '')
    add_input('Expansion Rate Pro (Monthly %) - Y1', monthly_expansion_rate['Pro'][1], 'Monthly % increase in revenue per retained user')
    add_input('Expansion Rate Pro (Monthly %) - Y2', monthly_expansion_rate['Pro'][13], '')
    add_input('Expansion Rate Pro (Monthly %) - Y3', monthly_expansion_rate['Pro'][25], '')
    add_input('Expansion Rate Team (Monthly %) - Y1', monthly_expansion_rate['Team'][1], 'Driven by users per account growth / cross-sell')
    add_input('Expansion Rate Team (Monthly %) - Y2', monthly_expansion_rate['Team'][13], '')
    add_input('Expansion Rate Team (Monthly %) - Y3', monthly_expansion_rate['Team'][25], '')
    add_input('Expansion Rate Enterprise (Monthly %) - Y1', monthly_expansion_rate['Enterprise'][1], 'Driven by users per account growth / cross-sell / modules')
    add_input('Expansion Rate Enterprise (Monthly %) - Y2', monthly_expansion_rate['Enterprise'][13], '')
    add_input('Expansion Rate Enterprise (Monthly %) - Y3', monthly_expansion_rate['Enterprise'][25], '')


    add_input('--- Account Assumptions ---', '', '')
    add_input('Avg Users per Team Account - Y1', avg_users_per_team_account[1], '')
    add_input('Avg Users per Team Account - Y2', avg_users_per_team_account[13], '')
    add_input('Avg Users per Team Account - Y3', avg_users_per_team_account[25], '')
    add_input('Avg Users per Enterprise Account - Y1', avg_users_per_enterprise_account[1], '')
    add_input('Avg Users per Enterprise Account - Y2', avg_users_per_enterprise_account[13], '')
    add_input('Avg Users per Enterprise Account - Y3', avg_users_per_enterprise_account[25], '')


    add_input('--- Pricing ---', '', '')
    add_input('Pro Monthly Price (€)', monthly_prices['Pro'][1], 'Fixed price per user')
    add_input('Team Monthly Price (€) - Y1', monthly_prices['Team'][1], 'Price per user per month')
    add_input('Team Monthly Price (€) - Y3', monthly_prices['Team'][25], 'Price per user per month') # Only Y1 and Y3 changed
    add_input('Enterprise Annual Price (€) - Y1', annual_prices['Enterprise'][1], 'Price per user per year')
    add_input('Enterprise Annual Price (€) - Y2', annual_prices['Enterprise'][13], '')
    add_input('Enterprise Annual Price (€) - Y3', annual_prices['Enterprise'][25], '')
    add_input('One-time License Price (€)', one_time_license_price[1], 'Paid by users converting to Lifetime tier')

    add_input('--- Marketplace Revenue (Calculated Annually) ---', '', '')
    add_input('Marketplace Commission Rate (%) - Y2', marketplace_commission_rate[13], 'Commission on purchases')
    add_input('Marketplace Commission Rate (%) - Y3', marketplace_commission_rate[25], '')
    add_input('Avg Marketplace Purchase (€/paid user/year) - Y2', avg_marketplace_purchase_per_user_per_year[13], 'Average spending per paid user per year')
    add_input('Avg Marketplace Purchase (€/paid user/year) - Y3', avg_marketplace_purchase_per_user_per_year[25], '')


    add_input('--- Cost Structure (% of Total Revenue) ---', '', '')
    add_input('Engineering Cost (%) - Y1', cost_structure_pct['Engineering'][1], 'Applied to Total Monthly Revenue')
    add_input('Engineering Cost (%) - Y2', cost_structure_pct['Engineering'][13], '')
    add_input('Engineering Cost (%) - Y3', cost_structure_pct['Engineering'][25], '')
    add_input('Product & Design Cost (%) - Y1', cost_structure_pct['Product & Design'][1], '')
    add_input('Product & Design Cost (%) - Y2', cost_structure_pct['Product & Design'][13], '')
    add_input('Product & Design Cost (%) - Y3', cost_structure_pct['Product & Design'][25], '')
    add_input('Marketing & Community Cost (%) - Y1', cost_structure_pct['Marketing & Community'][1], '')
    add_input('Marketing & Community Cost (%) - Y2', cost_structure_pct['Marketing & Community'][13], '')
    add_input('Marketing & Community Cost (%) - Y3', cost_structure_pct['Marketing & Community'][25], '')
    add_input('Customer Success Cost (%) - Y1', cost_structure_pct['Customer Success'][1], '')
    add_input('Customer Success Cost (%) - Y2', cost_structure_pct['Customer Success'][13], '')
    add_input('Customer Success Cost (%) - Y3', cost_structure_pct['Customer Success'][25], '')
    add_input('Infrastructure Cost (%) - Y1', cost_structure_pct['Infrastructure'][1], '')
    add_input('Infrastructure Cost (%) - Y2', cost_structure_pct['Infrastructure'][13], '')
    add_input('Infrastructure Cost (%) - Y3', cost_structure_pct['Infrastructure'][25], '')
    add_input('G&A Cost (%) - Y1', cost_structure_pct['G&A'][1], '')
    add_input('G&A Cost (%) - Y2', cost_structure_pct['G&A'][13], '')
    add_input('G&A Cost (%) - Y3', cost_structure_pct['G&A'][25], '')

    df_inputs = pd.DataFrame(df_inputs_data)
    df_inputs.to_excel(writer, sheet_name='Inputs', index=False)
    sheet = writer.sheets['Inputs']
    apply_styles(sheet, df_inputs) # Apply general styles

    # --- Onglet Monthly Model ---
    df_monthly.to_excel(writer, sheet_name='Monthly Model', index=False)
    sheet = writer.sheets['Monthly Model']
    apply_styles(sheet, df_monthly) # Apply general styles

    # --- Onglet Annual Summary ---
    df_annual.to_excel(writer, sheet_name='Annual Summary', index=True)
    sheet = writer.sheets['Annual Summary']
    apply_styles(sheet, df_annual) # Apply general styles


# --- Instruction Finale ---
print(f"Modèle financier THE ARCHITECT ECOSYSTEM créé : {output_filename}")
print("Ouvrez ce fichier Excel et ajustez les hypothèses dans l'onglet 'Inputs'.")
print("Les onglets 'Monthly Model' et 'Annual Summary' affichent les résultats calculés.")

