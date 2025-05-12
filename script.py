import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def creer_planificateur_weekend(nom_fichier="Planificateur_Weekend.xlsx"):
    # Créer un nouveau classeur et sélectionner la feuille active
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planificateur Weekend"
    
    # Définir les en-têtes
    ws['A1'] = "Nom"
    ws['B1'] = "Groupe"  # Nouvelle colonne pour les groupes
    ws['C1'] = "Transport"
    
    # Créer les en-têtes jour/heure
    jours = ["Mer", "Jeu", "Ven", "Sam", "Dim"]
    horaires = ["Matin", "Après-midi", "Soir"]
    
    col = 4  # Décalé d'une colonne à cause de la nouvelle colonne "Groupe"
    for jour in jours:
        for horaire in horaires:
            ws.cell(row=1, column=col, value=f"{jour} {horaire}")
            col += 1
    
    # Formater les en-têtes
    police_entete = Font(bold=True)
    for cellule in ws[1]:
        cellule.font = police_entete
        cellule.alignment = Alignment(horizontal='center')
    
    # Ajouter la validation des données pour Oui/Non/Peut-être
    validation_oui_non_peutetre = DataValidation(
        type="list",
        formula1='"Oui,Non,Peut-être"',
        allow_blank=True
    )
    ws.add_data_validation(validation_oui_non_peutetre)
    validation_oui_non_peutetre.add(f"D2:R20")  # Décalé à cause de la nouvelle colonne
    
    # Ajouter la validation des données pour le transport
    validation_transport = DataValidation(
        type="list",
        formula1='"Voiture,Train,Bus,Autre"',
        allow_blank=True
    )
    ws.add_data_validation(validation_transport)
    validation_transport.add(f"C2:C20")
    
    # Ajouter la validation pour les groupes (permettant des entrées libres mais suggérant des exemples)
    validation_groupe = DataValidation(
        type="list",
        formula1='"Groupe1,Groupe2,Groupe3,Voiture1,Voiture2,Train1"',
        allow_blank=True
    )
    ws.add_data_validation(validation_groupe)
    validation_groupe.add(f"B2:B20")
    
    # Créer une feuille pour la gestion des groupes
    ws_groupes = wb.create_sheet(title="Groupes")
    
    # Configurer la feuille de groupes
    ws_groupes['A1'] = "Nom du Groupe"
    ws_groupes['B1'] = "Description"
    ws_groupes['C1'] = "Membres"
    
    # Formatage des en-têtes de la feuille des groupes
    for cellule in ws_groupes[1]:
        cellule.font = police_entete
        cellule.alignment = Alignment(horizontal='center')
    
    # Exemples de groupes pré-remplis
    ws_groupes['A2'] = "Voiture1"
    ws_groupes['B2'] = "Personnes voyageant dans la Voiture 1"
    
    ws_groupes['A3'] = "Voiture2"
    ws_groupes['B3'] = "Personnes voyageant dans la Voiture 2"
    
    ws_groupes['A4'] = "Train1"
    ws_groupes['B4'] = "Personnes prenant le même train"
    
    # Ajuster automatiquement la largeur des colonnes (approximatif)
    for col in range(1, 19):  # Une colonne de plus
        lettre_colonne = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[lettre_colonne].width = 14
    
    # Ajuster les colonnes de la feuille des groupes
    ws_groupes.column_dimensions['A'].width = 15
    ws_groupes.column_dimensions['B'].width = 40
    ws_groupes.column_dimensions['C'].width = 40
    
    # Ajouter des instructions sur la première feuille
    ws.cell(row=22, column=1, value="Instructions:")
    ws.cell(row=23, column=1, value="- Utilisez la colonne 'Groupe' pour regrouper les personnes qui voyagent ensemble")
    ws.cell(row=24, column=1, value="- Vous pouvez créer et gérer vos groupes dans l'onglet 'Groupes'")
    ws.cell(row=25, column=1, value="- Les personnes du même groupe devraient généralement avoir des disponibilités similaires")
    
    # Bordures pour les instructions
    bordure_fine = Side(style='thin', color='000000')
    bordure = Border(top=bordure_fine, left=bordure_fine, right=bordure_fine, bottom=bordure_fine)
    
    for row in range(22, 26):
        cell = ws.cell(row=row, column=1)
        cell.border = bordure
    
    # Fusionner les cellules pour les instructions
    for row in range(22, 26):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    
    # Enregistrer le classeur
    wb.save(nom_fichier)
    print(f"Fichier planificateur de weekend créé: {nom_fichier}")
    print("Remarque: Veuillez ajouter manuellement le formatage conditionnel dans Excel:")
    print("- Formater les cellules contenant 'Oui' avec un fond vert")
    print("- Formater les cellules contenant 'Peut-être' avec un fond jaune")
    print("- Formater les cellules contenant 'Non' avec un fond rouge")
    print("\nNouvelle fonctionnalité de groupe ajoutée:")
    print("- Utilisez la colonne 'Groupe' pour identifier les personnes qui voyagent ensemble")
    print("- Consultez l'onglet 'Groupes' pour gérer les différents groupes")

if __name__ == "__main__":
    creer_planificateur_weekend()
